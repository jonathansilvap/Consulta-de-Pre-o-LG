from openpyxl import load_workbook
import pyautogui
import time
import pyperclip
from tkinter import *
from winotify import Notification

#planilha com os dados
planilha = load_workbook('D:\Automação Preço LG\AmbienteVirtual\Planilha de preco.xlsx')
aba_ativa = planilha['PREÇO'] # aba com os preços

#mudar para tela de apos play no script
pyautogui.keyDown('alt')
pyautogui.press(['tab'])
pyautogui.keyUp('alt')

#mover até o campo Cod peça
pyautogui.move(-1200,215)
#pyautogui.move(320,270)
pyautogui.click()
pyautogui.click()
linha_cell = 2
loc = -65

#obter os parts que serão consultados, serão consultados um a um
for linha in aba_ativa.iter_rows(min_row=2, max_row=False):
    pyautogui.write(linha[0].value)
    #time.sleep(1)

    #mover e cliar no campo de confirmação de dado
    pyautogui.move(0,33)
    time.sleep(2)
    pyautogui.click()

    #mover e selecionar o preço
    pyautogui.move(165,75)
    pyautogui.drag(xOffset=-110,yOffset=0,duration=0.5)
    #time.sleep(1)

    #copiar o preço para area de transferencia e salvar na variavel valor
    pyautogui.keyDown('ctrl')
    pyautogui.press(['c'])
    pyautogui.keyUp('ctrl')
    valor = pyperclip.paste()
       
    aba_ativa.cell(row=linha_cell,column=2,value= str(valor))

    linha_cell = linha_cell + 1
  
    #planilha.save(filename ='D:\Automação Preço LG\AmbienteVirtual\Planilha de preco v2.xlsx')
    #notificacao = Notification(app_id="Aviso", title="Notificação", msg="Arquivo .xlsx Pronto!!!", duration="short")

    #notificacao.show()

    
    pyautogui.moveTo(320,270)
    pyautogui.click()
    pyautogui.click()
    

planilha.save(filename ='D:\Automação Preço LG\AmbienteVirtual\Planilha de preco v2.xlsx')

notificacao = Notification(app_id="Aviso", title="Notificação", msg="Arquivo .xlsx Pronto!!!", duration="long")

notificacao.show()



#mudar para tela de apos play no script
#pyautogui.keyDown('alt')
#pyautogui.press(['tab'])
#pyautogui.keyUp('alt')

#aviso 1
#avs = Tk()0
#avs.title("Aviso")
#avs.geometry("300x100")

#lb_aviso = Label(avs,text="Favor aguardar, estou gerando o arquivo .xlsx")
#lb_aviso.place(x=10, y=-10, width=300, height=100)

#btn = Button(avs, text="OK", command=avs.after(5000, lambda: avs.destroy())).pack

#btn.place(x=135, y=60, width=30, height=30)

#avs.mainloop()

#planilha.save(filename ='D:\Automação Preço LG\AmbienteVirtual\Planilha de preco v2.xlsx')

#aviso 2

#avs2 = Tk()
#avs2.title("Aviso")
#avs2.geometry("300x100")

#lb_aviso2 = Label(avs2,text="Arquivo Salvo!")
#lb_aviso2.place(x=10, y=-10, width=300, height=100)

#btn2 = Button(avs2, text="OK", command=avs2.destroy)
#btn2.pack()
#btn2.place(x=135, y=60, width=30, height=30)

#avs2.mainloop()

#exit()
