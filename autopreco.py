from openpyxl import load_workbook
import pyautogui
import time
import pyperclip
from tkinter import *
from tkinter.filedialog import askdirectory
from winotify import Notification
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from PySimpleGUI import PySimpleGUI as sg


#layout da janela
sg.theme('Reddit')
layout = [
	[sg.Text('Senha Atual: motor6046***')],
    [sg.Text('Insira a Senha!'),sg.Input(key = 'senha',size=(20,1))],
 	[sg.Button('Entrar')]
]

#janela
# janela = sg.Window('Tela de Senha', layout)
# eventos, valores = janela.read()
# if eventos == sg.WINDOW_CLOSED:
# 	exit()
# if eventos == 'Entrar':
# 	if not valores['senha']:
# 		exit()
# 	else:
# 		senha = valores['senha']
# 		print(senha)
    
#planilha com os dados
#arq_origem = askdirectory(title = "Pasta Origem")
#arq_destino = askdirectory(title = "Pasta Destino")
planilha = load_workbook ('D:\Automação Preço LG\AmbienteVirtual\Planilha de preco.xlsx')
aba_ativa = planilha['PREÇO'] # aba com os preços

#login e senha
usuario = 'BR002199'
password = 'ruza6046*' #senha

#iniciando o navegador
service = Service(ChromeDriverManager().install())
nav = webdriver.Chrome(service=service)
nav.maximize_window()

#link do site
nav.get('https://saclge.com.br/svcportal/oow/requests/create')

#login no site
nav.find_element('id', 'username').send_keys(usuario)
nav.find_element('id','password').send_keys(password)
nav.find_element('xpath','//*[@id="landing-choice"]/div/div/form/div/div[2]/div[1]/center/button').click()
#time.sleep(10)

# arq_img1 = pyautogui.locateOnScreen('D:\Automação Preço LG\AmbienteVirtual\Cod_peça.PNG')

# if arq_img1 == None:
# 	layout2 = [	[sg.Text('SENHA INCORRETA')] ]
# 	janela2 = sg.Window('AVISO', layout2)
# 	evento = janela2.read()
# 	if evento == sg.WINDOW_CLOSED:
# 		exit()
# 	else:
# 		exit()
	
#coordenadar do campo de partnumber
pyautogui.moveTo(311,310)
pyautogui.click()
pyautogui.click()
linha_cell = 2

#obter os parts que serão consultados, serão consultados um a um
for linha in aba_ativa.iter_rows(min_row=2):
    if linha[0] is not None:
        pyautogui.write(linha[0].value)
        
        #mover e cliar no campo de confirmação de dado
        pyautogui.move(0,33)
        time.sleep(1.5)
        pyautogui.click()

        #mover e selecionar o preço
        pyautogui.move(165,75)
        pyautogui.drag(xOffset=-110,yOffset=0,duration=0.5)

        #copiar o preço para area de transferencia e salvar na variavel valor
        pyautogui.keyDown('ctrl')
        pyautogui.press(['c'])
        pyautogui.keyUp('ctrl')
        valor = pyperclip.paste()
            
        aba_ativa.cell(row=linha_cell,column=2,value= str(valor))

        linha_cell = linha_cell + 1
         
        pyautogui.moveTo(311,310)
        pyautogui.click()
        pyautogui.click()
    else:
        print('vazio') 

#salvar nova planilha        
planilha.save(filename ='D:\Automação Preço LG\AmbienteVirtual\Planilha de preco v2.xlsx')

#notificação de arquivo pronto
notificacao = Notification(app_id="Aviso", title="Notificação", msg="Arquivo .xlsx Pronto!!!", duration="long")
notificacao.show()